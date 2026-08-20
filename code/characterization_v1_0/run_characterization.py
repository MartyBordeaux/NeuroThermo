#!/usr/bin/env python3
import argparse, json, os, zipfile, tempfile, shutil
from pathlib import Path
import numpy as np
import pandas as pd
from scipy.stats import spearmanr
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

VERSION='1.0.0'

PARAMS=['b','r','s','kappa_I','kappa_over_Cm']
FIT_METRICS=['cell_loss','median_circle_f1','median_count_error_fraction','median_abs_latency_shift_ms']


def cliffs_delta(x,y):
    x=np.asarray(pd.Series(x).dropna(), dtype=float)
    y=np.asarray(pd.Series(y).dropna(), dtype=float)
    if len(x)==0 or len(y)==0:
        return np.nan
    gt=sum(a>b for a in x for b in y)
    lt=sum(a<b for a in x for b in y)
    return (gt-lt)/(len(x)*len(y))


def q25(s): return float(np.nanpercentile(np.asarray(s,dtype=float),25))
def q75(s): return float(np.nanpercentile(np.asarray(s,dtype=float),75))


def locate_results(root: Path):
    if root.is_file() and root.suffix.lower()=='.zip':
        td=Path(tempfile.mkdtemp(prefix='neurothermo_char_'))
        with zipfile.ZipFile(root) as z:
            z.extractall(td)
        candidates=list(td.rglob('cell_fit_summary.csv'))
        if len(candidates)!=1:
            raise RuntimeError(f'Expected exactly one cell_fit_summary.csv in zip; found {len(candidates)}')
        return candidates[0].parent, td
    if root.is_dir():
        if (root/'cell_fit_summary.csv').exists(): return root, None
        candidates=list(root.rglob('cell_fit_summary.csv'))
        if len(candidates)==1: return candidates[0].parent, None
    raise RuntimeError('Could not locate v3.9 result directory')


def phenotype_table(sw):
    rows=[]
    for cell_id,g in sw.groupby('cell_id',sort=True):
        g=g.sort_values('current_pA').copy()
        i_max=g['n_exp_spikes'].idxmax()
        maxrow=g.loc[i_max]
        last=g.iloc[-1]
        max_sp=float(maxrow['n_exp_spikes'])
        retention=float(last['n_exp_spikes']/max_sp) if max_sp>0 else np.nan
        rows.append({
            'cell_id':cell_id,
            'n_spiking_sweeps':int(len(g)),
            'max_exp_spikes':int(max_sp),
            'current_at_max_rate_pA':float(maxrow['current_pA']),
            'J_at_max_rate':float(maxrow['J']),
            'highest_spiking_current_pA':float(last['current_pA']),
            'exp_spikes_at_highest_current':int(last['n_exp_spikes']),
            'high_current_retention':retention,
            'depol_block_index':1.0-retention if np.isfinite(retention) else np.nan,
            'median_exp_latency_ms':float(g['exp_latency_ms'].median()),
            'min_exp_latency_ms':float(g['exp_latency_ms'].min()),
            'median_exp_train_duration_ms':float(g['exp_train_duration_ms'].median()),
            'median_abs_last_spike_error_ms':float(g['aligned_last_spike_error_ms'].abs().median()),
            'median_vp_loss_sweep':float(g['vp_loss'].median()),
        })
    return pd.DataFrame(rows)


def build_master(cf, sw, amap):
    accepted=cf[cf['final_v3_9_decision'].eq('ACCEPT')].copy()
    accepted['analysis_set']=np.where(accepted['primary_support'].eq('MULTI_SWEEP'),
                                      'PRIMARY_MULTI_SWEEP','SECONDARY_SINGLE_SWEEP')
    cols=['group','cell_id','animal_id','animal_id_status','experiment_day_code']
    accepted=accepted.merge(amap[cols],on=['group','cell_id'],how='left',validate='one_to_one')
    accepted['animal_id']=accepted['animal_id'].fillna('UNMAPPED')
    accepted['animal_resolved']=~accepted['animal_id'].isin(['NA_NOT_RECOVERABLE','UNMAPPED'])
    accepted['rheobase_pA']=accepted['threshold_first_spiking_current_pA']
    accepted['rheobase_J']=accepted['rheobase_pA']/accepted['capacitance_pF']
    ph=phenotype_table(sw)
    accepted=accepted.merge(ph,on='cell_id',how='left',validate='one_to_one')
    return accepted


def group_descriptive(primary, variables):
    rows=[]
    for group,g in primary.groupby('group',sort=True):
        for v in variables:
            x=pd.to_numeric(g[v],errors='coerce').dropna()
            if len(x)==0: continue
            rows.append({'level':'CELL','group':group,'variable':v,'n':len(x),
                         'median':float(x.median()),'q25':q25(x),'q75':q75(x),
                         'min':float(x.min()),'max':float(x.max())})
    return pd.DataFrame(rows)


def animal_table(primary):
    known=primary[primary['animal_resolved']].copy()
    numeric_cols=['capacitance_pF','cell_loss','median_circle_f1','median_count_error_fraction',
                  'b','r','s','kappa_I','kappa_over_Cm','rheobase_pA','rheobase_J',
                  'max_exp_spikes','depol_block_index','median_exp_latency_ms',
                  'median_exp_train_duration_ms','median_abs_last_spike_error_ms']
    agg=known.groupby(['group','animal_id'],as_index=False)[numeric_cols].median()
    counts=known.groupby(['group','animal_id']).size().rename('n_primary_cells').reset_index()
    return agg.merge(counts,on=['group','animal_id'])


def animal_group_descriptive(anim, variables):
    rows=[]
    for group,g in anim.groupby('group',sort=True):
        for v in variables:
            x=pd.to_numeric(g[v],errors='coerce').dropna()
            if len(x)==0: continue
            rows.append({'level':'ANIMAL_MEDIAN','group':group,'variable':v,'n_animals':len(x),
                         'median_of_animal_medians':float(x.median()),
                         'min_animal_median':float(x.min()),'max_animal_median':float(x.max())})
    return pd.DataFrame(rows)


def effect_sizes(primary, animals, variables):
    rows=[]
    for v in variables:
        xs=pd.to_numeric(primary.loc[primary.group=='SCA3',v],errors='coerce').dropna()
        xw=pd.to_numeric(primary.loc[primary.group=='WT',v],errors='coerce').dropna()
        if len(xs)==0 or len(xw)==0: continue
        as_=pd.to_numeric(animals.loc[animals.group=='SCA3',v],errors='coerce').dropna()
        aw=pd.to_numeric(animals.loc[animals.group=='WT',v],errors='coerce').dropna()
        separation='INSUFFICIENT_ANIMALS'
        if len(as_)>0 and len(aw)>0:
            if as_.min()>aw.max(): separation='SCA3_GT_WT_NO_OVERLAP'
            elif as_.max()<aw.min(): separation='SCA3_LT_WT_NO_OVERLAP'
            else: separation='ANIMAL_MEDIANS_OVERLAP'
        rows.append({
            'variable':v,
            'n_sca3_cells':len(xs),'n_wt_cells':len(xw),
            'median_sca3_cells':float(xs.median()),'median_wt_cells':float(xw.median()),
            'median_difference_sca3_minus_wt':float(xs.median()-xw.median()),
            'median_ratio_sca3_over_wt':float(xs.median()/xw.median()) if xw.median()!=0 else np.nan,
            'cliffs_delta_cell_level_descriptive':float(cliffs_delta(xs,xw)),
            'n_sca3_animals':len(as_),'n_wt_animals':len(aw),
            'median_sca3_animal_medians':float(as_.median()) if len(as_) else np.nan,
            'median_wt_animal_medians':float(aw.median()) if len(aw) else np.nan,
            'animal_median_ratio_sca3_over_wt':float(as_.median()/aw.median()) if len(as_) and len(aw) and aw.median()!=0 else np.nan,
            'animal_median_directional_separation':separation,
            'inference_note':'DESCRIPTIVE_ONLY: animal n is insufficient for reliable population inference'
        })
    return pd.DataFrame(rows)


def id_summary(primary):
    rows=[]
    cols=[('all_four','identifiability'),('b','id_b'),('r','id_r'),('s','id_s'),('kappa_I','id_kappa_I')]
    for scope,g in [('ALL',primary),('WT',primary[primary.group=='WT']),('SCA3',primary[primary.group=='SCA3'])]:
        for label,col in cols:
            n=int(g[col].notna().sum())
            ni=int(g[col].eq('IDENTIFIABLE').sum())
            rows.append({'scope':scope,'parameter':label,'n_evaluable':n,'n_identifiable':ni,
                         'fraction_identifiable':ni/n if n else np.nan})
    return pd.DataFrame(rows)


def correlations(primary):
    vars_=['b','r','s','kappa_I','kappa_over_Cm','capacitance_pF','rheobase_J',
           'max_exp_spikes','depol_block_index','median_exp_latency_ms','cell_loss',
           'median_circle_f1','median_count_error_fraction']
    rows=[]
    for scope,g in [('ALL',primary),('WT',primary[primary.group=='WT']),('SCA3',primary[primary.group=='SCA3'])]:
        for i,a in enumerate(vars_):
            for b in vars_[i+1:]:
                z=g[[a,b]].apply(pd.to_numeric,errors='coerce').dropna()
                if len(z)<4: continue
                rho,p=spearmanr(z[a],z[b])
                rows.append({'scope':scope,'var1':a,'var2':b,'n':len(z),'spearman_rho':rho,
                             'p_value_exploratory_not_inferential':p})
    return pd.DataFrame(rows)


def cohort_audit(master):
    rows=[]
    for aset,g in master.groupby('analysis_set'):
        for group,h in g.groupby('group'):
            known=h[h.animal_resolved]
            rows.append({'analysis_set':aset,'group':group,'n_cells':len(h),
                         'n_animal_resolved':int(known.shape[0]),
                         'n_animal_unresolved':int((~h.animal_resolved).sum()),
                         'n_distinct_recovered_animals':int(known.animal_id.nunique())})
    return pd.DataFrame(rows)


def write_excel(path, tables):
    with pd.ExcelWriter(path,engine='openpyxl') as xw:
        for name,df in tables.items():
            df.to_excel(xw,sheet_name=name[:31],index=False)
        wb=xw.book
        for ws in wb.worksheets:
            ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
            for cell in ws[1]:
                cell.font=Font(bold=True)
                cell.fill=PatternFill('solid',fgColor='D9EAF7')
                cell.alignment=Alignment(horizontal='center')
            for col in ws.columns:
                letter=get_column_letter(col[0].column)
                width=min(max(len(str(c.value)) if c.value is not None else 0 for c in col)+2,45)
                ws.column_dimensions[letter].width=max(width,10)


def jitter_positions(n,center):
    if n<=1:return np.array([center])
    return center+np.linspace(-0.07,0.07,n)


def plot_group_points(primary,var,ylabel,outpath,logy=False):
    fig,ax=plt.subplots(figsize=(4.8,5.2))
    groups=['WT','SCA3']
    for i,grp in enumerate(groups):
        vals=pd.to_numeric(primary.loc[primary.group==grp,var],errors='coerce').dropna().values
        ax.scatter(jitter_positions(len(vals),i),vals,s=40,alpha=.75)
        if len(vals):
            med=np.median(vals); lo=np.percentile(vals,25); hi=np.percentile(vals,75)
            ax.errorbar(i,med,yerr=[[med-lo],[hi-med]],fmt='s',ms=8,capsize=5,lw=2)
    ax.set_xticks([0,1],groups); ax.set_ylabel(ylabel)
    ax.set_title(f'Primary multi-sweep cells: {ylabel}')
    if logy: ax.set_yscale('log')
    fig.tight_layout(); fig.savefig(outpath,dpi=200); plt.close(fig)


def plot_animal_medians(anim,var,ylabel,outpath,logy=False):
    fig,ax=plt.subplots(figsize=(5.0,5.2))
    for i,grp in enumerate(['WT','SCA3']):
        g=anim[anim.group==grp]
        vals=pd.to_numeric(g[var],errors='coerce').values
        ax.scatter(jitter_positions(len(vals),i),vals,s=100,marker='D')
        for x,(_,r) in zip(jitter_positions(len(vals),i),g.iterrows()):
            ax.text(x+0.025,r[var],str(r.animal_id),fontsize=8,va='center')
    ax.set_xticks([0,1],['WT','SCA3']); ax.set_ylabel(f'Per-animal median {ylabel}')
    ax.set_title('Recovered animals only; descriptive')
    if logy: ax.set_yscale('log')
    fig.tight_layout(); fig.savefig(outpath,dpi=200); plt.close(fig)


def plot_kappa_cm(primary,outpath):
    fig,ax=plt.subplots(figsize=(5.5,4.8))
    for grp,mark in [('WT','o'),('SCA3','s')]:
        g=primary[primary.group==grp]
        ax.scatter(g.capacitance_pF,g.kappa_I,label=grp,marker=mark,s=55)
    ax.set_xscale('log'); ax.set_yscale('log')
    ax.set_xlabel('Capacitance (pF)'); ax.set_ylabel('kappa_I'); ax.legend()
    rho,_=spearmanr(primary.capacitance_pF,primary.kappa_I)
    ax.set_title(f'Input scaling vs capacitance; Spearman rho={rho:.2f}')
    fig.tight_layout(); fig.savefig(outpath,dpi=200); plt.close(fig)


def plot_id_matrix(primary,outpath):
    cols=['id_b','id_r','id_s','id_kappa_I']
    d=primary.sort_values(['group','cell_id']).copy()
    arr=np.array([[1 if r[c]=='IDENTIFIABLE' else 0 for c in cols] for _,r in d.iterrows()])
    fig,ax=plt.subplots(figsize=(5.5,max(5,0.28*len(d))))
    im=ax.imshow(arr,aspect='auto',vmin=0,vmax=1)
    ax.set_xticks(range(4),['b','r','s','kappa_I'])
    ax.set_yticks(range(len(d)),d['cell_id'])
    ax.set_title('Practical identifiability in v3.9')
    fig.colorbar(im,ax=ax,ticks=[0,1],label='0 non-ID, 1 ID')
    fig.tight_layout(); fig.savefig(outpath,dpi=200); plt.close(fig)


def write_report(out, primary, master, animals, effects, ids, audit):
    n_sca3=int((primary.group=='SCA3').sum()); n_wt=int((primary.group=='WT').sum())
    pa=audit[audit.analysis_set=='PRIMARY_MULTI_SWEEP'].set_index('group')
    def eff(v): return effects[effects.variable==v].iloc[0]
    lines=[]
    lines.append('# NeuroThermo post-fit characterization v1.0\n')
    lines.append('## Cohort and statistical unit\n')
    lines.append(f'Primary multi-sweep cohort: {len(primary)} cells ({n_wt} WT, {n_sca3} SCA3). The accepted fit cohort additionally contains {len(master)-len(primary)} single-sweep cells, retained only as secondary descriptive records.\n')
    lines.append(f"Recovered animal IDs in the primary cohort: WT {int(pa.loc['WT','n_animal_resolved'])}/{int(pa.loc['WT','n_cells'])} cells from {int(pa.loc['WT','n_distinct_recovered_animals'])} known animals; SCA3 {int(pa.loc['SCA3','n_animal_resolved'])}/{int(pa.loc['SCA3','n_cells'])} cells from {int(pa.loc['SCA3','n_distinct_recovered_animals'])} known animals.\n")
    lines.append('Because only two recovered animals are available per group, formal genotype-level population inference is not statistically reliable. Therefore the pipeline deliberately suppresses animal-level p-values. Cell-level group contrasts are reported as descriptive effect sizes for this recorded-cell ensemble, not as independent-animal inference.\n')
    lines.append('## Main parameter pattern\n')
    for v,label in [('b','b'),('r','r'),('s','s'),('kappa_I','kappa_I'),('kappa_over_Cm','kappa_I/Cm')]:
        e=eff(v)
        lines.append(f"- {label}: cell medians SCA3={e['median_sca3_cells']:.6g}, WT={e['median_wt_cells']:.6g}; SCA3/WT={e['median_ratio_sca3_over_wt']:.3f}; cell-level Cliff delta={e['cliffs_delta_cell_level_descriptive']:.3f}; recovered-animal pattern={e['animal_median_directional_separation']}.\n")
    lines.append('\nThe recovered-animal medians show complete directional separation for b and s (both SCA3 animals above both WT animals), and for raw kappa_I and capacitance (both SCA3 animals below both WT animals). There is no recovered-animal separation for r or kappa_I/Cm. This supports b and especially s as the more stable group-associated model coordinates, while r remains heterogeneous and raw kappa_I remains strongly entangled with capacitance.\n')
    lines.append('## Identifiability\n')
    iall=ids[ids.scope=='ALL'].set_index('parameter')
    lines.append(f"Full four-parameter identifiability: {int(iall.loc['all_four','n_identifiable'])}/{int(iall.loc['all_four','n_evaluable'])}. By parameter: b {int(iall.loc['b','n_identifiable'])}/{int(iall.loc['b','n_evaluable'])}, r {int(iall.loc['r','n_identifiable'])}/{int(iall.loc['r','n_evaluable'])}, s {int(iall.loc['s','n_identifiable'])}/{int(iall.loc['s','n_evaluable'])}, kappa_I {int(iall.loc['kappa_I','n_identifiable'])}/{int(iall.loc['kappa_I','n_evaluable'])}. Thus r remains the least identifiable coordinate.\n")
    lines.append('## Fit quality and phenotype descriptors\n')
    for v,label in [('cell_loss','cell loss'),('median_circle_f1','circle F1'),('median_count_error_fraction','count-error fraction'),('capacitance_pF','capacitance'),('rheobase_J','rheobase current density')]:
        e=eff(v)
        lines.append(f"- {label}: SCA3 median={e['median_sca3_cells']:.6g}, WT median={e['median_wt_cells']:.6g}, SCA3/WT={e['median_ratio_sca3_over_wt']:.3f}.\n")
    lines.append('\nSCA3 fits are systematically harder at the cell level (higher loss/count error and lower timing F1). This must be shown alongside parameter comparisons so that group differences are not interpreted independently of model fit quality.\n')
    lines.append('## Publication interpretation\n')
    lines.append('The current dataset supports rigorous cell-level dynamical characterization and animal-aware descriptive comparisons. It does not support strong animal-population significance claims, because the recovered SCA3 cohort originates from two animals. For a manuscript, biological group claims should therefore be phrased as patterns within the recorded-cell ensemble and supported by per-animal medians. Strong population-level genotype inference would require additional independent animals or recovery of additional SCA3 animal identities demonstrating a larger animal count.\n')
    (out/'REPORT.md').write_text(''.join(lines),encoding='utf-8')


def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--results',required=True,help='v3.9 results zip or directory')
    ap.add_argument('--animal-map',required=True,help='NeuroThermo_animal_id_recovery.xlsx')
    ap.add_argument('--outdir',required=True)
    args=ap.parse_args()
    out=Path(args.outdir); out.mkdir(parents=True,exist_ok=True)
    rdir,tmp=locate_results(Path(args.results))
    try:
        cf=pd.read_csv(rdir/'cell_fit_summary.csv')
        sw=pd.read_csv(rdir/'sweep_fit_summary.csv')
        amap=pd.read_excel(args.animal_map,sheet_name='Accepted cohort')
        master=build_master(cf,sw,amap)
        primary=master[master.analysis_set=='PRIMARY_MULTI_SWEEP'].copy()
        animals=animal_table(primary)
        variables=['b','r','s','kappa_I','kappa_over_Cm','capacitance_pF','cell_loss','median_circle_f1',
                   'median_count_error_fraction','rheobase_pA','rheobase_J','max_exp_spikes','depol_block_index',
                   'median_exp_latency_ms','median_exp_train_duration_ms','median_abs_last_spike_error_ms']
        gd=group_descriptive(primary,variables)
        ad=animal_group_descriptive(animals,variables)
        effects=effect_sizes(primary,animals,variables)
        ids=id_summary(primary)
        cors=correlations(primary)
        audit=cohort_audit(master)

        master.to_csv(out/'cell_level_master.csv',index=False)
        animals.to_csv(out/'animal_level_medians_recovered.csv',index=False)
        gd.to_csv(out/'group_descriptive_cell_level.csv',index=False)
        ad.to_csv(out/'group_descriptive_animal_level.csv',index=False)
        effects.to_csv(out/'group_effect_sizes_descriptive.csv',index=False)
        ids.to_csv(out/'identifiability_summary.csv',index=False)
        cors.to_csv(out/'spearman_correlations_descriptive.csv',index=False)
        audit.to_csv(out/'cohort_animal_id_audit.csv',index=False)

        policy={
            'version':VERSION,
            'primary_analysis_set':'PRIMARY_MULTI_SWEEP',
            'primary_cells':int(len(primary)),
            'primary_wt_cells':int((primary.group=='WT').sum()),
            'primary_sca3_cells':int((primary.group=='SCA3').sum()),
            'recovered_animals_per_group':animals.groupby('group').animal_id.nunique().to_dict(),
            'formal_animal_level_pvalues':'SUPPRESSED',
            'reason':'Only two recovered animals per group; cell observations are nested within animals and are not independent biological replicates.',
            'cell_level_effect_sizes':'DESCRIPTIVE_ONLY',
            'single_sweep_cells':'SECONDARY_DESCRIPTIVE_ONLY',
            'special_cell_sensitivity_analysis':'NONE'
        }
        (out/'STATISTICAL_GUARDRAILS.json').write_text(json.dumps(policy,indent=2),encoding='utf-8')

        figdir=out/'figures'; figdir.mkdir(exist_ok=True)
        specs=[('b','b',False),('r','r',True),('s','s',False),('kappa_I','kappa_I',True),
               ('kappa_over_Cm','kappa_I / Cm',True),('capacitance_pF','Capacitance (pF)',False),
               ('cell_loss','Cell loss',False),('rheobase_J','Rheobase J (pA/pF)',True)]
        for v,label,logy in specs:
            plot_group_points(primary,v,label,figdir/f'cell_{v}.png',logy)
        for v,label,logy in [('b','b',False),('r','r',True),('s','s',False),('kappa_I','kappa_I',True),
                             ('kappa_over_Cm','kappa_I / Cm',True),('capacitance_pF','Capacitance (pF)',False)]:
            plot_animal_medians(animals,v,label,figdir/f'animal_{v}.png',logy)
        plot_kappa_cm(primary,figdir/'kappa_vs_capacitance.png')
        plot_id_matrix(primary,figdir/'identifiability_matrix.png')

        tables={
            'Cohort audit':audit,
            'Cell master':master,
            'Animal medians':animals,
            'Cell group summary':gd,
            'Animal group summary':ad,
            'Effect sizes':effects,
            'Identifiability':ids,
            'Correlations':cors,
        }
        write_excel(out/'NeuroThermo_characterization_summary.xlsx',tables)
        write_report(out,primary,master,animals,effects,ids,audit)

        summary={
            'version':VERSION,
            'n_primary_cells':len(primary),
            'primary_by_group':primary.groupby('group').size().to_dict(),
            'recovered_primary_cells_by_group':primary[primary.animal_resolved].groupby('group').size().to_dict(),
            'recovered_animals_by_group':animals.groupby('group').animal_id.nunique().to_dict(),
            'full_identifiable':int(primary.identifiability.eq('IDENTIFIABLE').sum()),
            'id_b':int(primary.id_b.eq('IDENTIFIABLE').sum()),
            'id_r':int(primary.id_r.eq('IDENTIFIABLE').sum()),
            'id_s':int(primary.id_s.eq('IDENTIFIABLE').sum()),
            'id_kappa_I':int(primary.id_kappa_I.eq('IDENTIFIABLE').sum()),
        }
        (out/'SUMMARY.json').write_text(json.dumps(summary,indent=2),encoding='utf-8')
        print(json.dumps(summary,indent=2))
    finally:
        if tmp is not None: shutil.rmtree(tmp,ignore_errors=True)

if __name__=='__main__': main()
